import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
import io
import os

import boto3
from botocore.exceptions import ClientError

from fastapi import FastAPI, Request, Form, HTTPException, Response, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from .db import (
    put_match, get_match, list_matches, list_matches_by_user, update_match, delete_match,
    list_matches_by_tournament,
    put_event, list_events, delete_last_event,
    create_user, get_user_by_username, get_user_by_id, get_user_by_email,
    list_all_users, delete_user, update_user_password, toggle_user_active,
    set_user_role, search_users, update_user_stats, set_user_must_change_password,
    set_user_admin,
    put_tournament, get_tournament, list_tournaments, list_tournaments_by_user,
    list_tournaments_for_scorer,
    update_tournament, delete_tournament,
    get_settings, update_settings,
    add_roster_player, list_roster, delete_roster_player, get_roster_player,
    put_registration, get_registration, list_registrations_by_tournament,
    list_all_registrations, update_registration_paid, delete_registration,
    find_registration_by_name,
)
from .logic import compute_state, player_name
from .auth import hash_password, verify_password, create_session_token, verify_session_token

app = FastAPI(title="Table Tennis Scorer")
templates = Jinja2Templates(directory="app/templates")
app.mount("/static", StaticFiles(directory="app/static"), name="static")


# ── S3 photo storage ─────────────────────────────────────────────────────────
# Photos live in a private S3 bucket (TT_S3_BUCKET). We store the S3 key on
# each registration and generate a short-lived presigned URL when rendering.
AWS_REGION = os.getenv("AWS_REGION", "eu-west-1")
S3_BUCKET = os.getenv("TT_S3_BUCKET", "").strip()
S3_PREFIX = os.getenv("TT_S3_PREFIX", "registrations").strip("/")
PRESIGN_TTL_SECONDS = int(os.getenv("TT_S3_PRESIGN_TTL", "3600"))
ALLOWED_PHOTO_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
ALLOWED_PHOTO_MIME = {
    "image/jpeg", "image/png", "image/webp", "image/gif",
}
MAX_PHOTO_BYTES = 5 * 1024 * 1024  # 5 MB
s3_client = boto3.client("s3", region_name=AWS_REGION)


def _upload_photo_to_s3(upload: "UploadFile", registration_id: str) -> str:
    """Store an uploaded photo in S3 under registrations/{id}.{ext}.
    Returns the S3 key. Raises HTTPException on bad input / missing bucket."""
    if not S3_BUCKET:
        raise HTTPException(500, "Photo uploads disabled — TT_S3_BUCKET not configured")
    filename = (upload.filename or "").lower()
    ext = "." + filename.rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in ALLOWED_PHOTO_EXT:
        raise HTTPException(400, f"Photo must be one of {sorted(ALLOWED_PHOTO_EXT)}")
    content_type = (upload.content_type or "").lower()
    if content_type and content_type not in ALLOWED_PHOTO_MIME:
        raise HTTPException(400, f"Unsupported photo content-type: {content_type}")
    data = upload.file.read(MAX_PHOTO_BYTES + 1)
    if len(data) > MAX_PHOTO_BYTES:
        raise HTTPException(400, f"Photo too large (max {MAX_PHOTO_BYTES // (1024*1024)} MB)")
    if not data:
        raise HTTPException(400, "Empty photo upload")
    key = f"{S3_PREFIX}/{registration_id}{ext}"
    try:
        s3_client.put_object(
            Bucket=S3_BUCKET,
            Key=key,
            Body=data,
            ContentType=content_type or "application/octet-stream",
            ServerSideEncryption="AES256",
        )
    except ClientError as e:
        raise HTTPException(500, f"S3 upload failed: {e.response.get('Error', {}).get('Code', 'Unknown')}")
    return key


def _delete_photo_from_s3(key: str) -> None:
    if not key or not S3_BUCKET:
        return
    try:
        s3_client.delete_object(Bucket=S3_BUCKET, Key=key)
    except ClientError:
        pass  # best-effort


def _presigned_photo_url(key: str) -> Optional[str]:
    """Short-lived signed GET URL for private S3 objects. Empty key → None."""
    if not key or not S3_BUCKET:
        return None
    try:
        return s3_client.generate_presigned_url(
            "get_object",
            Params={"Bucket": S3_BUCKET, "Key": key},
            ExpiresIn=PRESIGN_TTL_SECONDS,
        )
    except ClientError:
        return None


templates.env.globals["photo_url"] = _presigned_photo_url


def _advancing_participants(tournament: dict, round_num: int) -> list:
    """Participants eligible to be picked in round `round_num`.
    Round 1 → all participants.
    Round N ≥ 2 → only the winners of round N‑1's completed matches.
        Singles winner: 1 participant (the winning slot).
        Doubles winner: 2 participants (both members of the winning pair).
    Falls back to all participants if the previous round has no decided winners
    yet (so the dropdown is never empty and losers can still be manually chosen
    for byes / corrections)."""
    participants = tournament.get("participants", []) or []
    if round_num <= 1:
        return participants
    rounds = tournament.get("rounds", []) or []
    prev = next((r for r in rounds if int(r.get("round_num", 0)) == int(round_num) - 1), None)
    if not prev:
        return participants
    advancing_ids = set()
    for m in prev.get("matches", []) or []:
        w = m.get("winner")  # "A" or "B" once the match is scored
        if not w:
            continue
        if w == "A":
            advancing_ids.add(m.get("a_id"))
            if m.get("match_type") == "doubles" and m.get("a2_id"):
                advancing_ids.add(m.get("a2_id"))
        elif w == "B":
            advancing_ids.add(m.get("b_id"))
            if m.get("match_type") == "doubles" and m.get("b2_id"):
                advancing_ids.add(m.get("b2_id"))
    if not advancing_ids:
        return participants  # nothing decided yet → don't lock the UI
    return [p for p in participants if p.get("id") in advancing_ids]


templates.env.globals["advancing_participants"] = _advancing_participants


def _winning_side_label(match: Dict[str, Any], tournament: Optional[Dict[str, Any]],
                        winner: str) -> str:
    """Human label for the winning side of a match.
    Singles → the winning player's name.
    Doubles → the team name (from the tournament's participants) if we can
    resolve it, otherwise "Primary & Partner". Falls back to the primary name
    if we somehow have no partner info at all."""
    if winner not in ("A", "B"):
        return ""
    if winner == "A":
        primary = (match.get("player_a") or "").strip()
        partner = (match.get("player_a2") or "").strip()
    else:
        primary = (match.get("player_b") or "").strip()
        partner = (match.get("player_b2") or "").strip()
    if (match.get("match_type") or "singles") != "doubles":
        return primary
    # Doubles: prefer team_name recorded on either member's participant entry.
    team = ""
    if tournament:
        by_name = {(p.get("name") or "").strip().lower(): p
                   for p in (tournament.get("participants") or [])}
        for nm in (primary, partner):
            if not nm:
                continue
            p = by_name.get(nm.lower())
            if p and p.get("team_name"):
                team = p["team_name"]
                break
    if team:
        return team
    if primary and partner:
        return f"{primary} & {partner}"
    return primary or partner


def _slot_winner_label(tournament: Optional[Dict[str, Any]], slot: Dict[str, Any]) -> str:
    """Render-time helper for bracket slots. Prefers the tournament team_name
    (via participants lookup by id) over the stored winner_name so historical
    matches (built before doubles team labelling) still show a good label."""
    if not slot:
        return ""
    winner = slot.get("winner")
    if winner not in ("A", "B"):
        return slot.get("winner_name") or ""
    participants = {p.get("id"): p for p in (tournament.get("participants") or [])} if tournament else {}
    if winner == "A":
        primary_id = slot.get("a_id")
        partner_id = slot.get("a2_id")
    else:
        primary_id = slot.get("b_id")
        partner_id = slot.get("b2_id")
    primary = participants.get(primary_id) or {}
    partner = participants.get(partner_id) or {}
    if (slot.get("match_type") or "singles") != "doubles":
        return (primary.get("name") or slot.get("winner_name") or "").strip()
    team = (primary.get("team_name") or partner.get("team_name") or "").strip()
    if team:
        return team
    pn = (primary.get("name") or "").strip()
    qn = (partner.get("name") or primary.get("partner_name") or "").strip()
    if pn and qn:
        return f"{pn} & {qn}"
    return pn or slot.get("winner_name") or ""


templates.env.globals["winner_slot_label"] = _slot_winner_label


@app.get("/sw.js")
async def service_worker():
    return FileResponse(
        "app/static/sw.js",
        media_type="application/javascript",
        headers={"Service-Worker-Allowed": "/"},
    )


@app.get("/favicon.ico")
async def favicon():
    return FileResponse("app/static/icons/icon-192.png", media_type="image/png")


@app.exception_handler(404)
async def not_found_handler(request: Request, exc: HTTPException):
    return templates.TemplateResponse("404.html", {"request": request}, status_code=404)


class ForcePasswordChange(Exception):
    """Raised when an authenticated user must change their password before continuing."""


@app.exception_handler(ForcePasswordChange)
async def force_password_change_handler(request: Request, exc: ForcePasswordChange):
    return RedirectResponse("/change-password?forced=1", status_code=303)


# ── Helpers ────────────────────────────────────────────────────────────────────
def now_ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ") + "#" + uuid.uuid4().hex


def get_current_user(request: Request) -> Optional[Dict[str, Any]]:
    token = request.cookies.get("session")
    if not token:
        return None
    user_id = verify_session_token(token)
    if not user_id:
        return None
    return get_user_by_id(user_id)


def require_auth(request: Request) -> Dict[str, Any]:
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    # If admin has forced a password reset, block every route except the change-password flow itself.
    if user.get("must_change_password"):
        path = request.url.path
        if not (path.startswith("/change-password") or path == "/logout"):
            raise ForcePasswordChange()
    return user


def require_admin(request: Request) -> Dict[str, Any]:
    user = require_auth(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


def _is_root_admin(user: Dict[str, Any]) -> bool:
    """Root admin = the seeded 'admin' account. Only role that can mint new admins."""
    if not user or not user.get("is_admin"):
        return False
    return bool(user.get("is_root_admin")) or user.get("username") == "admin"


def require_root_admin(request: Request) -> Dict[str, Any]:
    user = require_admin(request)
    if not _is_root_admin(user):
        raise HTTPException(status_code=403, detail="Only the root admin can manage admins")
    return user


def _render_user_management(request: Request, **extra):
    """Render the user-management partial with the current user's root-admin flag."""
    current_user = get_current_user(request)
    ctx = {
        "request": request,
        "all_users": list_all_users(),
        "is_root_admin": _is_root_admin(current_user or {}),
    }
    ctx.update(extra)
    return templates.TemplateResponse("partials/user_management.html", ctx)


def require_scorer(request: Request) -> Dict[str, Any]:
    user = require_auth(request)
    if user.get("role", "scorer") == "player" and not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Scorer access required")
    return user


def must_match(match_id: str) -> Dict[str, Any]:
    m = get_match(match_id)
    if not m:
        raise HTTPException(status_code=404, detail="Match not found")
    return m


def must_tournament(tournament_id: str) -> Dict[str, Any]:
    t = get_tournament(tournament_id)
    if not t:
        raise HTTPException(status_code=404, detail="Tournament not found")
    return t


def _tournament_grants_scorer(t: Dict[str, Any], user: Dict[str, Any]) -> bool:
    """True if the user is listed as a scorer on the tournament."""
    if not t or not user:
        return False
    return user.get("user_id", "") in (t.get("scorer_ids") or [])


def check_match_access(request: Request, match_id: str):
    user = require_auth(request)
    match = must_match(match_id)
    if user.get("is_admin"):
        return user, match
    if match.get("user_id") == user["user_id"]:
        return user, match
    # Scorers assigned to the parent tournament may score its matches.
    tid = match.get("tournament_id") or ""
    if tid:
        t = get_tournament(tid)
        if t and _tournament_grants_scorer(t, user):
            return user, match
    raise HTTPException(status_code=403, detail="Access denied")


def check_tournament_access(request: Request, tournament_id: str):
    user = require_auth(request)
    t = must_tournament(tournament_id)
    if user.get("is_admin"):
        return user, t
    if t.get("user_id") == user["user_id"]:
        return user, t
    if _tournament_grants_scorer(t, user):
        return user, t
    raise HTTPException(status_code=403, detail="Access denied")


def match_context(request: Request, match: Dict[str, Any], user: Optional[Dict[str, Any]] = None,
                  flash: Optional[str] = None) -> Dict[str, Any]:
    ev = list_events(match["match_id"])
    state = compute_state(match, ev)
    tid = match.get("tournament_id", "")
    photos: Dict[str, Optional[str]] = {"a": None, "b": None, "a2": None, "b2": None}
    team_a = team_b = ""
    if tid:
        # Photos for each slot
        for slot, field in (("a", "player_a"), ("b", "player_b"),
                            ("a2", "player_a2"), ("b2", "player_b2")):
            nm = (match.get(field) or "").strip()
            if not nm:
                continue
            reg = find_registration_by_name(tid, nm)
            if reg and reg.get("photo_key"):
                photos[slot] = _presigned_photo_url(reg["photo_key"])
        # Team names (doubles): resolve from either primary or partner registration
        if match.get("match_type") == "doubles":
            for side, primary_field, partner_field in (
                ("a", "player_a", "player_a2"),
                ("b", "player_b", "player_b2"),
            ):
                nm = (match.get(primary_field) or "").strip()
                if not nm:
                    continue
                reg = find_registration_by_name(tid, nm)
                team = (reg.get("team_name") if reg else "") or ""
                if not team:
                    # try the partner side too
                    nm2 = (match.get(partner_field) or "").strip()
                    if nm2:
                        reg2 = find_registration_by_name(tid, nm2)
                        team = (reg2.get("team_name") if reg2 else "") or ""
                if side == "a":
                    team_a = team
                else:
                    team_b = team
    return {
        "request": request,
        "user": user,
        "match": match,
        "state": state,
        "flash": flash,
        "photo_a": photos["a"],
        "photo_b": photos["b"],
        "photo_a2": photos["a2"],
        "photo_b2": photos["b2"],
        "team_a": team_a,
        "team_b": team_b,
    }


# ── Auth ──────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    if user.get("must_change_password"):
        return RedirectResponse("/change-password?forced=1", status_code=303)
    if user.get("role", "scorer") == "player" and not user.get("is_admin"):
        return RedirectResponse(f"/profile/{user['user_id']}", status_code=303)

    if user.get("is_admin"):
        recent_matches = list_matches(limit=50)
        recent_tournaments = list_tournaments(limit=20)
    else:
        recent_matches = list_matches_by_user(user["user_id"], limit=50)
        # Scorers see tournaments they created AND ones admins assigned them to.
        recent_tournaments = list_tournaments_for_scorer(user["user_id"], limit=20)

    # Only admins get the scorer picker on the create form.
    all_scorers = []
    if user.get("is_admin"):
        all_scorers = [
            {"user_id": u["user_id"], "username": u.get("username", "")}
            for u in list_all_users()
            if u.get("is_active") and not u.get("is_admin")
            and (u.get("role") or "scorer") != "player"
        ]
        all_scorers.sort(key=lambda u: u["username"].lower())

    return templates.TemplateResponse("home.html", {
        "request": request,
        "user": user,
        "recent_matches": recent_matches,
        "recent_tournaments": recent_tournaments,
        "settings": get_settings(),
        "all_scorers": all_scorers,
    })


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if get_current_user(request):
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    user = get_user_by_username(username)
    if not user or not verify_password(password, user["password_hash"]):
        return templates.TemplateResponse("login.html", {
            "request": request, "error": "Invalid credentials"
        }, status_code=401)

    if not user.get("is_active", True):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Your account has not been activated yet. Contact the admin."
        }, status_code=403)

    token = create_session_token(user["user_id"])
    if user.get("must_change_password"):
        redirect_to = "/change-password?forced=1"
    elif user.get("role", "scorer") == "player" and not user.get("is_admin"):
        redirect_to = f"/profile/{user['user_id']}"
    else:
        redirect_to = "/"
    response = RedirectResponse(redirect_to, status_code=303)
    response.set_cookie(key="session", value=token, httponly=True, max_age=86400 * 7)
    return response


@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request):
    if get_current_user(request):
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("register.html", {"request": request})


@app.post("/register")
async def register(request: Request, username: str = Form(...), email: str = Form(...),
                   password: str = Form(...), role: str = Form("scorer")):
    if get_user_by_username(username):
        return templates.TemplateResponse("register.html", {
            "request": request, "error": "Username already exists"
        }, status_code=400)
    if get_user_by_email(email):
        return templates.TemplateResponse("register.html", {
            "request": request, "error": "Email already registered"
        }, status_code=400)

    # Auth sign-up is scorer-only. Players register per-tournament via the
    # public /tournaments/{id}/register form (no account required).
    role = "scorer"

    user_id = uuid.uuid4().hex
    create_user(user_id, username, hash_password(password), is_admin=False, email=email, role=role)
    success_msg = "Scorer account created. Contact the admin to activate your account before logging in."

    return templates.TemplateResponse("register.html", {
        "request": request, "success": success_msg
    })


@app.get("/logout")
def logout():
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie("session")
    return response


@app.get("/change-password", response_class=HTMLResponse)
def change_password_page(request: Request, forced: int = 0):
    user = require_auth(request)
    return templates.TemplateResponse("change_password.html", {
        "request": request, "user": user,
        "forced": bool(forced) or bool(user.get("must_change_password")),
    })


@app.post("/change-password")
def change_password(request: Request,
                    current_password: str = Form(...),
                    new_password: str = Form(...),
                    confirm_password: str = Form(...)):
    user = require_auth(request)
    forced = bool(user.get("must_change_password"))
    ctx = {"request": request, "user": user, "forced": forced}
    if not verify_password(current_password, user["password_hash"]):
        return templates.TemplateResponse("change_password.html", {**ctx, "error": "Current password is incorrect"})
    if new_password != confirm_password:
        return templates.TemplateResponse("change_password.html", {**ctx, "error": "New passwords do not match"})
    if len(new_password) < 6:
        return templates.TemplateResponse("change_password.html", {**ctx, "error": "Password must be at least 6 characters"})
    if new_password == current_password:
        return templates.TemplateResponse("change_password.html", {**ctx, "error": "New password must differ from the current one"})
    update_user_password(user["user_id"], hash_password(new_password))
    if forced:
        set_user_must_change_password(user["user_id"], False)
        # Send them home now that the wall is lifted.
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("change_password.html", {
        **ctx, "forced": False, "success": "Password updated successfully!"
    })


# ── User search ───────────────────────────────────────────────────────────────
@app.get("/users/search")
def users_search(request: Request, q: str = "", exclude: str = ""):
    require_auth(request)
    if len(q) < 2:
        return JSONResponse([])
    exclude_ids = [e for e in exclude.split(",") if e]
    return JSONResponse(search_users(q, exclude_ids=exclude_ids))


# ── Profile ───────────────────────────────────────────────────────────────────
@app.get("/profile/{user_id}", response_class=HTMLResponse)
def profile_page(request: Request, user_id: str):
    current_user = require_auth(request)
    profile_user = get_user_by_id(user_id)
    if not profile_user:
        raise HTTPException(404, "User not found")
    stats = {
        "matches_played": int(profile_user.get("stat_matches_played", 0)),
        "matches_won": int(profile_user.get("stat_matches_won", 0)),
        "games_played": int(profile_user.get("stat_games_played", 0)),
        "games_won": int(profile_user.get("stat_games_won", 0)),
        "points_scored": int(profile_user.get("stat_points_scored", 0)),
    }
    stats["match_win_rate"] = round(stats["matches_won"] / stats["matches_played"] * 100) if stats["matches_played"] else 0
    stats["game_win_rate"] = round(stats["games_won"] / stats["games_played"] * 100) if stats["games_played"] else 0
    return templates.TemplateResponse("profile.html", {
        "request": request,
        "current_user": current_user,
        "profile_user": profile_user,
        "stats": stats,
    })


# ── Admin ─────────────────────────────────────────────────────────────────────
@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request):
    admin = require_admin(request)
    all_matches = list_matches(limit=200)
    all_users = list_all_users()
    all_tournaments = list_tournaments(limit=50)
    return templates.TemplateResponse("admin.html", {
        "request": request, "user": admin,
        "settings": get_settings(),
        "roster": list_roster(),
        "matches": all_matches,
        "all_users": all_users,
        "tournaments": all_tournaments,
        "is_root_admin": _is_root_admin(admin),
    })


@app.post("/admin/settings", response_class=HTMLResponse)
async def admin_update_settings(request: Request,
                                 default_best_of: int = Form(5),
                                 default_points_to_win: int = Form(11),
                                 service_interval: int = Form(2),
                                 deuce_interval: int = Form(1),
                                 default_match_type: str = Form("singles"),
                                 deciding_side_change_at: int = Form(5),
                                 hard_cap_enabled: str = Form(""),
                                 hard_cap_at: int = Form(15)):
    require_admin(request)
    if default_best_of not in (1, 3, 5, 7):
        raise HTTPException(400, "default_best_of must be 1, 3, 5, or 7")
    if default_points_to_win < 5:
        raise HTTPException(400, "default_points_to_win must be at least 5")
    if service_interval < 1 or deuce_interval < 1:
        raise HTTPException(400, "intervals must be at least 1")
    if default_match_type not in ("singles", "doubles"):
        raise HTTPException(400, "default_match_type must be 'singles' or 'doubles'")
    if deciding_side_change_at < 0:
        raise HTTPException(400, "deciding_side_change_at must be >= 0 (0 disables)")
    if hard_cap_at < 5:
        raise HTTPException(400, "hard_cap_at must be at least 5")
    update_settings({
        "default_best_of": int(default_best_of),
        "default_points_to_win": int(default_points_to_win),
        "service_interval": int(service_interval),
        "deuce_interval": int(deuce_interval),
        "default_match_type": default_match_type,
        "deciding_side_change_at": int(deciding_side_change_at),
        "hard_cap_enabled": hard_cap_enabled in ("on", "true", "1", "yes"),
        "hard_cap_at": int(hard_cap_at),
    })
    return templates.TemplateResponse("partials/settings_form.html", {
        "request": request,
        "settings": get_settings(),
        "flash": "✅ Settings saved. New matches will use these defaults.",
    })


@app.post("/admin/roster", response_class=HTMLResponse)
async def admin_add_roster(request: Request, name: str = Form(...), user_id: str = Form("")):
    require_admin(request)
    nm = name.strip()
    if not nm:
        raise HTTPException(400, "Name required")
    pid = uuid.uuid4().hex[:10]
    add_roster_player(pid, nm, user_id=user_id.strip())
    return templates.TemplateResponse("partials/roster.html", {
        "request": request,
        "roster": list_roster(),
        "flash": f"✅ Added {nm} to roster",
    })


@app.delete("/admin/roster/{player_id}", response_class=HTMLResponse)
async def admin_delete_roster(request: Request, player_id: str):
    require_admin(request)
    delete_roster_player(player_id)
    return templates.TemplateResponse("partials/roster.html", {
        "request": request,
        "roster": list_roster(),
    })


@app.post("/admin/matches/{match_id}/delete")
async def admin_delete_match(request: Request, match_id: str):
    require_admin(request)
    delete_match(match_id)
    return HTMLResponse("", status_code=200, headers={"HX-Refresh": "true"})


@app.post("/admin/tournaments/{tournament_id}/delete")
async def admin_delete_tournament(request: Request, tournament_id: str):
    require_admin(request)
    _cascade_delete_tournament(tournament_id)
    return HTMLResponse("", status_code=200, headers={"HX-Refresh": "true"})


def _cascade_delete_tournament(tournament_id: str) -> Dict[str, int]:
    """Delete a tournament plus all of its registrations and matches.
    Returns a small stats dict for logging/UX."""
    regs = list_registrations_by_tournament(tournament_id)
    for r in regs:
        delete_registration(r["registration_id"])
    ms = list_matches_by_tournament(tournament_id)
    for m in ms:
        delete_match(m["match_id"])
    delete_tournament(tournament_id)
    return {"registrations": len(regs), "matches": len(ms)}


@app.post("/tournaments/{tournament_id}/delete", response_class=HTMLResponse)
def delete_tournament_route(request: Request, tournament_id: str):
    """Cascade-delete a tournament from its own page. Admin or owner only."""
    user, t = check_tournament_access(request, tournament_id)
    _cascade_delete_tournament(tournament_id)
    # After delete, send the browser home.
    return HTMLResponse("", status_code=200, headers={"HX-Redirect": "/"})


@app.post("/admin/users/{user_id}/toggle-active")
async def admin_toggle_user(request: Request, user_id: str):
    require_admin(request)
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")
    if target.get("is_admin"):
        raise HTTPException(400, "Cannot deactivate admin users")
    toggle_user_active(user_id, not target.get("is_active", True))
    return _render_user_management(request)


@app.post("/admin/users/{user_id}/make-scorer")
async def admin_make_scorer(request: Request, user_id: str):
    require_admin(request)
    set_user_role(user_id, "scorer")
    toggle_user_active(user_id, True)
    return _render_user_management(request)


@app.post("/admin/users/{user_id}/make-player")
async def admin_make_player(request: Request, user_id: str):
    require_admin(request)
    set_user_role(user_id, "player")
    toggle_user_active(user_id, True)
    return _render_user_management(request)


@app.post("/admin/users/{user_id}/reset-password")
async def admin_reset_password(request: Request, user_id: str):
    require_admin(request)
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")
    if target.get("is_admin"):
        raise HTTPException(400, "Cannot reset another admin's password from here")
    import secrets, string
    alphabet = string.ascii_letters + string.digits
    temp_password = ''.join(secrets.choice(alphabet) for _ in range(10))
    update_user_password(user_id, hash_password(temp_password))
    set_user_must_change_password(user_id, True)
    username = target["username"]
    return _render_user_management(
        request,
        reset_password_msg=(
            f"🔑 Temp password for {username}: {temp_password} "
            f"— share it securely. They must change it on next login."
        ),
    )


@app.post("/admin/users/create-admin", response_class=HTMLResponse)
async def admin_create_admin(request: Request,
                             username: str = Form(...),
                             password: str = Form(""),
                             email: str = Form("")):
    """Root-admin-only: create a brand-new admin account."""
    require_root_admin(request)
    uname = (username or "").strip().lower()
    if not uname:
        raise HTTPException(400, "Username is required")
    if get_user_by_username(uname):
        raise HTTPException(400, f"Username '{uname}' is already taken")

    pw = (password or "").strip()
    generated = False
    if not pw:
        import secrets, string
        alphabet = string.ascii_letters + string.digits
        pw = ''.join(secrets.choice(alphabet) for _ in range(12))
        generated = True
    elif len(pw) < 8:
        raise HTTPException(400, "Password must be at least 8 characters")

    new_id = uuid.uuid4().hex
    create_user(
        user_id=new_id,
        username=uname,
        password_hash=hash_password(pw),
        is_admin=True,
        email=(email or "").strip(),
        role="scorer",
    )
    # Force password change on first login when we generated the temp password.
    if generated:
        set_user_must_change_password(new_id, True)

    msg = (
        f"👑 Admin '{uname}' created. Temp password: {pw} — share securely, "
        f"they must change it on first login."
    ) if generated else f"👑 Admin '{uname}' created."
    return _render_user_management(request, reset_password_msg=msg)


@app.post("/admin/users/{user_id}/revoke-admin", response_class=HTMLResponse)
async def admin_revoke_admin(request: Request, user_id: str):
    """Root-admin-only: demote another admin back to scorer. Cannot revoke root."""
    root = require_root_admin(request)
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")
    if not target.get("is_admin"):
        raise HTTPException(400, "User is not an admin")
    if target.get("is_root_admin") or target.get("username") == "admin":
        raise HTTPException(400, "Cannot revoke the root admin")
    if target["user_id"] == root["user_id"]:
        raise HTTPException(400, "You cannot revoke your own admin rights")
    set_user_admin(user_id, False)
    set_user_role(user_id, "scorer")
    return _render_user_management(request)


@app.delete("/admin/users/{user_id}")
async def admin_delete_user(request: Request, user_id: str):
    require_admin(request)
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")
    if target.get("is_admin"):
        raise HTTPException(400, "Cannot delete admin users")
    # Cascade delete: this user's matches
    for m in list_matches_by_user(user_id):
        delete_match(m["match_id"])
    delete_user(user_id)
    return _render_user_management(request)


# ── Matches: create + view ────────────────────────────────────────────────────
@app.post("/matches")
def create_match(request: Request,
                 name: str = Form(...),
                 player_a: str = Form(...),
                 player_b: str = Form(...),
                 match_type: str = Form(""),
                 player_a2: str = Form(""),
                 player_b2: str = Form(""),
                 best_of: int = Form(0),
                 points_to_win: int = Form(0),
                 service_interval: int = Form(0),
                 deuce_interval: int = Form(0),
                 deciding_side_change_at: int = Form(-1),
                 first_server: str = Form("A"),
                 first_server_side: int = Form(1),
                 first_receiver_side: int = Form(1),
                 tournament_id: str = Form(""),
                 round_num: int = Form(0),
                 match_slot: int = Form(0)):
    user = require_scorer(request)
    cfg = get_settings()
    best_of = int(best_of) or int(cfg["default_best_of"])
    points_to_win = int(points_to_win) or int(cfg["default_points_to_win"])
    service_interval = int(service_interval) or int(cfg["service_interval"])
    deuce_interval = int(deuce_interval) or int(cfg["deuce_interval"])
    if deciding_side_change_at < 0:
        deciding_side_change_at = int(cfg.get("deciding_side_change_at", 5))
    # Snapshot the hard-cap rule onto the match so changing global settings
    # later won't retroactively alter finished/in-progress matches.
    hard_cap_enabled = bool(cfg.get("hard_cap_enabled", False))
    hard_cap_at = int(cfg.get("hard_cap_at", 15))
    if match_type not in ("singles", "doubles"):
        match_type = cfg.get("default_match_type", "singles")

    if best_of not in (1, 3, 5, 7):
        raise HTTPException(400, "best_of must be 1, 3, 5, or 7")
    if points_to_win < 5:
        raise HTTPException(400, "points_to_win must be at least 5")
    if service_interval < 1 or deuce_interval < 1:
        raise HTTPException(400, "intervals must be at least 1")
    if first_server not in ("A", "B"):
        first_server = "A"
    if first_server_side not in (1, 2):
        first_server_side = 1
    if first_receiver_side not in (1, 2):
        first_receiver_side = 1
    if match_type == "doubles":
        if not player_a2.strip() or not player_b2.strip():
            raise HTTPException(400, "Doubles requires two players per side")

    match_id = uuid.uuid4().hex
    item = {
        "match_id": match_id,
        "name": name.strip(),
        "match_type": match_type,
        "player_a": player_a.strip() or "Player A",
        "player_b": player_b.strip() or "Player B",
        "best_of": int(best_of),
        "points_to_win": int(points_to_win),
        "service_interval": int(service_interval),
        "deuce_interval": int(deuce_interval),
        "deciding_side_change_at": int(deciding_side_change_at),
        "hard_cap_enabled": hard_cap_enabled,
        "hard_cap_at": hard_cap_at,
        "first_server": first_server,
        "created_at": now_ts(),
        "user_id": user["user_id"],
        "stats_recorded": False,
    }
    if match_type == "doubles":
        item["player_a2"] = player_a2.strip()
        item["player_b2"] = player_b2.strip()
        item["first_server_side"] = int(first_server_side)
        item["first_receiver_side"] = int(first_receiver_side)
    if tournament_id:
        item["tournament_id"] = tournament_id
        item["round_num"] = int(round_num)
        item["match_slot"] = int(match_slot)
    put_match(item)

    if tournament_id:
        # link the new match_id to the tournament round/slot
        t = get_tournament(tournament_id)
        if t:
            rounds = t.get("rounds", [])
            for r in rounds:
                if int(r.get("round_num", 0)) == int(round_num):
                    for slot in r.get("matches", []):
                        if int(slot.get("slot", 0)) == int(match_slot):
                            slot["match_id"] = match_id
                            break
            update_tournament(tournament_id, "SET rounds = :r", {":r": rounds})

    return RedirectResponse(f"/matches/{match_id}", status_code=303)


@app.get("/matches/{match_id}", response_class=HTMLResponse)
def match_page(request: Request, match_id: str):
    user, match = check_match_access(request, match_id)
    ctx = match_context(request, match, user=user)
    return templates.TemplateResponse("match.html", ctx)


@app.get("/live", response_class=HTMLResponse)
def live_index(request: Request):
    """Public rolled-up board: every tournament grouped by round with all
    matches shown inline (scores + game state). No click-through required."""
    tournaments_view: List[Dict[str, Any]] = []
    standalone_live: List[Dict[str, Any]] = []
    seen_match_ids: set = set()

    for t in list_tournaments(limit=50):
        rounds_view = _build_rounds_view(t)
        # Track match ids that belong to a tournament (so the stray-matches
        # section below doesn't duplicate them).
        for r in rounds_view:
            for m in r["matches"]:
                if m.get("match_id"):
                    seen_match_ids.add(m["match_id"])
        has_matches = any(r["matches"] for r in rounds_view)
        if not has_matches:
            continue
        # Only surface tournaments that have at least one live or upcoming match.
        # (Keep finished ones too — they're informational and small.)
        tournaments_view.append({
            "tournament": t,
            "rounds_view": rounds_view,
            "live_count": sum(
                1 for r in rounds_view for m in r["matches"] if m["status"] == "live"
            ),
        })

    # Ad-hoc (non-tournament) live matches — keep the previous flat behaviour
    # so casual scoreboards still show up here.
    for m in list_matches(limit=200):
        if m.get("tournament_id"):
            continue
        if m["match_id"] in seen_match_ids:
            continue
        ev = list_events(m["match_id"])
        st = compute_state(m, ev)
        if st.get("match_winner"):
            continue
        standalone_live.append({
            "match_id": m["match_id"],
            "name": m.get("name", "Match"),
            "player_a": m.get("player_a", "A"),
            "player_b": m.get("player_b", "B"),
            "a_score": st.get("a_score", 0),
            "b_score": st.get("b_score", 0),
            "a_games": st.get("a_games", 0),
            "b_games": st.get("b_games", 0),
            "current_game_num": st.get("current_game_num", 1),
            "best_of": m.get("best_of", 5),
            "is_doubles": (m.get("match_type") == "doubles"),
        })
    standalone_live.sort(key=lambda r: r["name"].lower())

    # Most-recently-created tournament first.
    tournaments_view.sort(
        key=lambda tv: tv["tournament"].get("created_at", ""), reverse=True,
    )
    return templates.TemplateResponse("live_index.html", {
        "request": request,
        "tournaments_view": tournaments_view,
        "standalone_live": standalone_live,
    })


@app.get("/live/{match_id}", response_class=HTMLResponse)
def live_match(request: Request, match_id: str):
    match = must_match(match_id)
    ctx = match_context(request, match, user=None)
    return templates.TemplateResponse("live.html", ctx)


# ── Matches: scoring actions ──────────────────────────────────────────────────
def _finalize_stats_if_needed(match: dict, state: dict) -> None:
    if not state.get("match_winner") or match.get("stats_recorded"):
        return
    a_id = match.get("player_a_id")
    b_id = match.get("player_b_id")
    a_points = sum(g["a"] for g in state["games"])
    b_points = sum(g["b"] for g in state["games"])
    if a_id:
        update_user_stats(a_id,
                          match_won=(state["match_winner"] == "A"),
                          games_played=state["a_games"] + state["b_games"],
                          games_won=state["a_games"],
                          points_scored=a_points)
    if b_id:
        update_user_stats(b_id,
                          match_won=(state["match_winner"] == "B"),
                          games_played=state["a_games"] + state["b_games"],
                          games_won=state["b_games"],
                          points_scored=b_points)
    update_match(match["match_id"], "SET stats_recorded = :s, winner = :w",
                 {":s": True, ":w": state["match_winner"]})

    # If part of a tournament, record the winner in the tournament's bracket
    tid = match.get("tournament_id")
    if tid:
        t = get_tournament(tid)
        if t:
            rounds = t.get("rounds", [])
            for r in rounds:
                if int(r.get("round_num", 0)) == int(match.get("round_num", 0)):
                    for slot in r.get("matches", []):
                        if slot.get("match_id") == match["match_id"]:
                            slot["winner"] = state["match_winner"]
                            slot["winner_name"] = _winning_side_label(match, t, state["match_winner"])
                            break
            update_tournament(tid, "SET rounds = :r", {":r": rounds})


@app.post("/matches/{match_id}/point/{ab}", response_class=HTMLResponse)
def add_point(request: Request, match_id: str, ab: str):
    user, match = check_match_access(request, match_id)
    if ab not in ("A", "B"):
        raise HTTPException(400, "scorer must be A or B")

    ev = list_events(match_id)
    state = compute_state(match, ev)
    if state.get("match_winner"):
        return templates.TemplateResponse("partials/scoreboard.html", {
            "request": request, "user": user, "match": match, "state": state,
            "flash": "🏆 Match is already over."
        })

    put_event({
        "match_id": match_id,
        "ts": now_ts(),
        "scorer": ab,
        "game_num": state["current_game_num"],
        "undone": False,
    })
    match2 = must_match(match_id)
    state2 = compute_state(match2, list_events(match_id))
    _finalize_stats_if_needed(match2, state2)
    if state2.get("match_winner"):
        match2 = must_match(match_id)
        state2 = compute_state(match2, list_events(match_id))

    flash = None
    if state2.get("match_winner"):
        winner_nm = player_name(match2, state2["match_winner"])
        flash = f"🏆 Match Over! {winner_nm} wins!"
    elif state2.get("side_change_alert"):
        flash = "🔀 Change ends! (deciding game — swap sides)"
    elif state2.get("service_change"):
        server_nm = state2.get("server_name") or player_name(match2, state2["server"])
        flash = f"🔁 Service change — {server_nm} to serve"

    return templates.TemplateResponse("partials/scoreboard.html", {
        "request": request, "user": user, "match": match2, "state": state2, "flash": flash,
    })


@app.post("/matches/{match_id}/let", response_class=HTMLResponse)
def add_let(request: Request, match_id: str):
    """Record a let (net on serve) — no point awarded, serve is replayed."""
    user, match = check_match_access(request, match_id)
    state = compute_state(match, list_events(match_id))
    if state.get("match_winner"):
        return templates.TemplateResponse("partials/scoreboard.html", {
            "request": request, "user": user, "match": match, "state": state,
            "flash": "🏆 Match is already over."
        })
    put_event({
        "match_id": match_id,
        "ts": now_ts(),
        "type": "let",
        "scorer": None,
        "game_num": state["current_game_num"],
        "undone": False,
    })
    match2 = must_match(match_id)
    state2 = compute_state(match2, list_events(match_id))
    return templates.TemplateResponse("partials/scoreboard.html", {
        "request": request, "user": user, "match": match2, "state": state2,
        "flash": "🌐 Let — serve replayed (no point).",
    })


@app.post("/matches/{match_id}/undo", response_class=HTMLResponse)
def undo_point(request: Request, match_id: str):
    user, match = check_match_access(request, match_id)
    deleted = delete_last_event(match_id)
    match2 = must_match(match_id)
    state2 = compute_state(match2, list_events(match_id))
    flash = "↩️ Last point undone." if deleted else "⚠️ No points to undo."
    return templates.TemplateResponse("partials/scoreboard.html", {
        "request": request, "user": user, "match": match2, "state": state2, "flash": flash,
    })


@app.post("/matches/{match_id}/players", response_class=HTMLResponse)
def set_players(request: Request, match_id: str,
                player_a: str = Form(""), player_b: str = Form(""),
                player_a2: str = Form(""), player_b2: str = Form(""),
                player_a_id: str = Form(""), player_b_id: str = Form("")):
    user, match = check_match_access(request, match_id)
    updates = []
    vals: Dict[str, Any] = {}
    if player_a.strip():
        updates.append("player_a = :pa")
        vals[":pa"] = player_a.strip()
    if player_b.strip():
        updates.append("player_b = :pb")
        vals[":pb"] = player_b.strip()
    if player_a2.strip():
        updates.append("player_a2 = :pa2")
        vals[":pa2"] = player_a2.strip()
    if player_b2.strip():
        updates.append("player_b2 = :pb2")
        vals[":pb2"] = player_b2.strip()
    if player_a_id:
        updates.append("player_a_id = :paid")
        vals[":paid"] = player_a_id
    if player_b_id:
        updates.append("player_b_id = :pbid")
        vals[":pbid"] = player_b_id
    if updates:
        update_match(match_id, "SET " + ", ".join(updates), vals)
    match2 = must_match(match_id)
    state2 = compute_state(match2, list_events(match_id))
    return templates.TemplateResponse("partials/scoreboard.html", {
        "request": request, "user": user, "match": match2, "state": state2,
        "flash": "✅ Players updated",
    })


@app.post("/matches/{match_id}/delete")
def delete_own_match(request: Request, match_id: str):
    user, match = check_match_access(request, match_id)
    delete_match(match_id)
    return RedirectResponse("/", status_code=303)


# ── Tournaments ───────────────────────────────────────────────────────────────
@app.post("/tournaments")
async def create_tournament(request: Request,
                      name: str = Form(...),
                      best_of: int = Form(0),
                      points_to_win: int = Form(0),
                      service_interval: int = Form(0),
                      deuce_interval: int = Form(0),
                      rounds_seed: str = Form(""),
                      format: str = Form("doubles"),
                      registration_start: str = Form(""),
                      registration_end: str = Form("")):
    # Only admins can create tournaments.
    user = require_admin(request)
    form = await request.form()
    # Multi-select "scorer_ids" (0..N values); FastAPI's Form(...) only sees the
    # last one, so we read them off the raw form.
    scorer_ids = [v for v in form.getlist("scorer_ids") if v]
    cfg = get_settings()
    tid = uuid.uuid4().hex

    fmt = (format or "doubles").strip().lower()
    if fmt not in ("singles", "doubles"):
        fmt = "doubles"

    # Optional: seed rounds from a textarea. One round name per line
    # (commas and semicolons also accepted).
    seeded_rounds = []
    if rounds_seed.strip():
        raw = rounds_seed.replace(",", "\n").replace(";", "\n").splitlines()
        names = [n.strip() for n in raw if n.strip()]
        for idx, rname in enumerate(names, start=1):
            seeded_rounds.append({
                "round_num": idx,
                "name": rname,
                "matches": [],
            })

    put_tournament({
        "tournament_id": tid,
        "name": name.strip(),
        "best_of": int(best_of) or int(cfg["default_best_of"]),
        "points_to_win": int(points_to_win) or int(cfg["default_points_to_win"]),
        "service_interval": int(service_interval) or int(cfg["service_interval"]),
        "deuce_interval": int(deuce_interval) or int(cfg["deuce_interval"]),
        "user_id": user["user_id"],
        "scorer_ids": scorer_ids,
        "created_at": now_ts(),
        "participants": [],
        "rounds": seeded_rounds,
        "format": fmt,
        "registration_start": registration_start.strip(),
        "registration_end": registration_end.strip(),
        "status": "registration",
    })
    return RedirectResponse(f"/tournaments/{tid}", status_code=303)


@app.get("/tournaments/{tournament_id}", response_class=HTMLResponse)
def tournament_page(request: Request, tournament_id: str):
    user, t = check_tournament_access(request, tournament_id)
    all_scorers = []
    if user.get("is_admin"):
        all_scorers = [
            {"user_id": u["user_id"], "username": u.get("username", "")}
            for u in list_all_users()
            if u.get("is_active") and not u.get("is_admin")
            and (u.get("role") or "scorer") != "player"
        ]
        all_scorers.sort(key=lambda u: u["username"].lower())
    return templates.TemplateResponse("tournament.html", {
        "request": request, "user": user, "tournament": t,
        "roster": list_roster(),
        "all_scorers": all_scorers,
    })


@app.post("/tournaments/{tournament_id}/scorers", response_class=HTMLResponse)
async def update_tournament_scorers(request: Request, tournament_id: str):
    """Admin-only: replace the tournament's scorer_ids list."""
    user = require_admin(request)
    t = must_tournament(tournament_id)
    if t.get("user_id") != user["user_id"] and not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    form = await request.form()
    scorer_ids = [v for v in form.getlist("scorer_ids") if v]
    update_tournament(
        tournament_id,
        "SET scorer_ids = :s",
        {":s": scorer_ids},
    )
    return RedirectResponse(f"/tournaments/{tournament_id}", status_code=303)


# ── Public tournament registration (players sign up to play) ─────────────────
def _registration_window_state(t: Dict[str, Any]) -> Dict[str, Any]:
    """Return {is_open, reason} based on tournament registration window + status."""
    status = (t.get("status") or "registration").lower()
    if status != "registration":
        return {"is_open": False, "reason": "Registration is closed — the tournament has already started."}
    start_s = (t.get("registration_start") or "").strip()
    end_s = (t.get("registration_end") or "").strip()
    now = now_ts()  # ISO-8601 in UTC
    # datetime-local inputs are naive local time; we compare as strings which
    # is not perfectly accurate across timezones but is fine for open/close gating.
    if start_s and now < start_s:
        return {"is_open": False, "reason": f"Registration opens on {start_s.replace('T', ' ')[:16]}."}
    if end_s and now > end_s:
        return {"is_open": False, "reason": f"Registration closed on {end_s.replace('T', ' ')[:16]}."}
    return {"is_open": True, "reason": ""}


@app.get("/tournaments/{tournament_id}/register", response_class=HTMLResponse)
def registration_form(request: Request, tournament_id: str):
    t = must_tournament(tournament_id)  # no auth — anyone with the link
    window = _registration_window_state(t)
    # Admin/owner sees the form even if the window is closed (for manual add).
    current_user = get_current_user(request)
    is_admin_view = bool(current_user and (
        current_user.get("is_admin") or current_user.get("user_id") == t.get("user_id")
    ))
    if is_admin_view:
        window = {"is_open": True, "reason": ""}
    return templates.TemplateResponse("tournament_register.html", {
        "request": request, "tournament": t, "flash": "",
        "window": window,
        "is_admin_view": is_admin_view,
    })


@app.post("/tournaments/{tournament_id}/register", response_class=HTMLResponse)
async def registration_submit(request: Request, tournament_id: str,
                              name: str = Form(...),
                              email: str = Form(...),
                              phone: str = Form(""),
                              its: str = Form(""),
                              age: int = Form(0),
                              experience: str = Form("beginner"),
                              match_type: str = Form("singles"),
                              team_name: str = Form(""),
                              partner_name: str = Form(""),
                              partner_email: str = Form(""),
                              partner_phone: str = Form(""),
                              partner_its: str = Form(""),
                              partner_age: int = Form(0),
                              partner_experience: str = Form("beginner"),
                              photo: Optional[UploadFile] = File(None),
                              partner_photo: Optional[UploadFile] = File(None)):
    t = must_tournament(tournament_id)

    # Admins/owners can bypass the registration window (used by the manual-add flow).
    is_manual_add = False
    current_user = get_current_user(request)
    if current_user and (
        current_user.get("is_admin") or current_user.get("user_id") == t.get("user_id")
    ):
        is_manual_add = True

    if not is_manual_add:
        window = _registration_window_state(t)
        if not window["is_open"]:
            raise HTTPException(403, window["reason"])

    def _render_error(msg: str, status: int = 400):
        """Render the register page with a friendly error banner instead of a JSON 4xx.
        Called for user-facing validation / duplicate failures.
        """
        return templates.TemplateResponse(
            "tournament_register.html",
            {
                "request": request,
                "tournament": t,
                "is_admin_view": is_manual_add,
                "window": _registration_window_state(t),
                "flash_error": msg,
            },
            status_code=status,
        )

    name = name.strip()
    email = email.strip().lower()
    phone_v = phone.strip()
    its_v = its.strip()
    experience = experience.strip().lower()
    try:
        age_val = int(age) if age else 0
    except (TypeError, ValueError):
        age_val = 0

    # Primary block — every field is required
    missing_primary = []
    if not name: missing_primary.append("name")
    if not email: missing_primary.append("email")
    if not phone_v: missing_primary.append("phone")
    if not its_v: missing_primary.append("ITS number")
    if age_val <= 0: missing_primary.append("age")
    if experience not in ("beginner", "amateur", "expert"):
        missing_primary.append("experience")
    if photo is None or not (photo.filename or "").strip():
        missing_primary.append("photo")
    if missing_primary:
        return _render_error("Missing required fields: " + ", ".join(missing_primary), 400)

    match_type = (match_type or "singles").strip().lower()
    if match_type not in ("singles", "doubles"):
        match_type = "singles"

    # Validate partner block for doubles
    partner_name = partner_name.strip()
    partner_email = partner_email.strip().lower()
    partner_phone_v = partner_phone.strip()
    partner_its_v = partner_its.strip()
    partner_experience = partner_experience.strip().lower()
    try:
        partner_age_val = int(partner_age) if partner_age else 0
    except (TypeError, ValueError):
        partner_age_val = 0

    if match_type == "doubles":
        missing = []
        if not partner_name: missing.append("partner name")
        if not partner_email: missing.append("partner email")
        if not partner_phone_v: missing.append("partner phone")
        if not partner_its_v: missing.append("partner ITS number")
        if partner_age_val <= 0: missing.append("partner age")
        if partner_experience not in ("beginner", "amateur", "expert"):
            missing.append("partner experience")
        if partner_photo is None or not (partner_photo.filename or "").strip():
            missing.append("partner photo")
        if not (team_name or "").strip():
            missing.append("team name")
        if missing:
            return _render_error("Doubles registration missing: " + ", ".join(missing), 400)
        if partner_name.lower() == name.lower():
            return _render_error("Partner must be a different person.", 400)
    if partner_experience not in ("beginner", "amateur", "expert"):
        partner_experience = "beginner"

    team_name = (team_name or "").strip()

    # ── Duplicate check: no player may register twice in the same tournament.
    # We compare on any of {email, ITS, phone} (normalized). This catches both:
    #   • someone who already registered as singles trying to re-register
    #   • someone who was added as a partner in a doubles pair
    existing_regs = list_registrations_by_tournament(tournament_id)

    def _norm(s: str) -> str:
        return (s or "").strip().lower()

    def _norm_phone(s: str) -> str:
        # phones may include spaces / dashes / plus — compare digits only
        return "".join(ch for ch in (s or "") if ch.isdigit())

    taken_emails = {_norm(r.get("email", "")) for r in existing_regs if r.get("email")}
    taken_its = {_norm(r.get("its", "")) for r in existing_regs if r.get("its")}
    taken_phones = {_norm_phone(r.get("phone", "")) for r in existing_regs if r.get("phone")}

    def _who_conflicts(reg_email: str, reg_its: str, reg_phone: str) -> str:
        e = _norm(reg_email)
        i = _norm(reg_its)
        p = _norm_phone(reg_phone)
        if e and e in taken_emails:
            return f"email {reg_email}"
        if i and i in taken_its:
            return f"ITS {reg_its}"
        if p and p in taken_phones:
            return f"phone {reg_phone}"
        return ""

    primary_conflict = _who_conflicts(email, its_v, phone_v)
    if primary_conflict:
        return _render_error(
            f"{name} is already registered for this tournament ({primary_conflict}). "
            "If you registered as a partner in a doubles pair, please contact the organizer.",
            409,
        )
    if match_type == "doubles":
        partner_conflict = _who_conflicts(partner_email, partner_its_v, partner_phone_v)
        if partner_conflict:
            return _render_error(
                f"Your partner {partner_name} is already registered for this tournament "
                f"({partner_conflict}).",
                409,
            )
        # Also guard within THIS submission: primary and partner must not share any identifier.
        if _norm(email) and _norm(email) == _norm(partner_email):
            return _render_error("Partner email must be different from yours.", 400)
        if _norm(its_v) and _norm(its_v) == _norm(partner_its_v):
            return _render_error("Partner ITS must be different from yours.", 400)
        if _norm_phone(phone_v) and _norm_phone(phone_v) == _norm_phone(partner_phone_v):
            return _render_error("Partner phone must be different from yours.", 400)

    pair_id = uuid.uuid4().hex if match_type == "doubles" else ""

    # Primary registration
    reg_id = uuid.uuid4().hex
    photo_key = ""
    if photo is not None and (photo.filename or "").strip():
        photo_key = _upload_photo_to_s3(photo, reg_id)
    primary_item = {
        "registration_id": reg_id,
        "tournament_id": tournament_id,
        "name": name,
        "email": email,
        "phone": phone_v,
        "its": its_v,
        "age": age_val,
        "experience": experience,
        "photo_key": photo_key,
        "payment_done": False,
        "created_at": now_ts(),
        "match_type": match_type,
    }
    if match_type == "doubles":
        primary_item["pair_id"] = pair_id
        primary_item["partner_name"] = partner_name
        primary_item["team_name"] = team_name
    put_registration(primary_item)

    # Partner registration (linked by pair_id)
    if match_type == "doubles":
        partner_reg_id = uuid.uuid4().hex
        partner_photo_key = ""
        if partner_photo is not None and (partner_photo.filename or "").strip():
            partner_photo_key = _upload_photo_to_s3(partner_photo, partner_reg_id)
        put_registration({
            "registration_id": partner_reg_id,
            "tournament_id": tournament_id,
            "name": partner_name,
            "email": partner_email,
            "phone": partner_phone_v,
            "its": partner_its_v,
            "age": partner_age_val,
            "experience": partner_experience,
            "photo_key": partner_photo_key,
            "payment_done": False,
            "created_at": now_ts(),
            "match_type": "doubles",
            "pair_id": pair_id,
            "partner_name": name,
            "team_name": team_name,
        })

    flash_msg = (
        f"✅ Thanks {name}! Team \"{team_name}\" ({name} & {partner_name}) is registered for "
        f"{t.get('name', 'the tournament')}."
        if match_type == "doubles"
        else f"✅ Thanks {name}! You're registered for {t.get('name', 'the tournament')}. See you at the table."
    )
    return templates.TemplateResponse("tournament_register.html", {
        "request": request, "tournament": t,
        "flash": flash_msg,
        "just_registered": True,
    })


# ── Registrations management (scorer/admin) ──────────────────────────────────
@app.get("/tournaments/{tournament_id}/registrations", response_class=HTMLResponse)
def registrations_page(request: Request, tournament_id: str):
    user, t = check_tournament_access(request, tournament_id)
    regs = list_registrations_by_tournament(tournament_id)
    return templates.TemplateResponse("tournament_registrations.html", {
        "request": request, "user": user, "tournament": t, "registrations": regs,
    })


@app.post("/registrations/{registration_id}/toggle-paid", response_class=HTMLResponse)
def registrations_toggle_paid(request: Request, registration_id: str):
    reg = get_registration(registration_id)
    if not reg:
        raise HTTPException(404, "Registration not found")
    # Only admin or the tournament owner may flip payment status
    user, t = check_tournament_access(request, reg["tournament_id"])
    update_registration_paid(registration_id, not bool(reg.get("payment_done")))
    return HTMLResponse("", status_code=200, headers={"HX-Refresh": "true"})


@app.post("/registrations/{registration_id}/delete", response_class=HTMLResponse)
def registrations_delete(request: Request, registration_id: str):
    reg = get_registration(registration_id)
    if not reg:
        raise HTTPException(404, "Registration not found")
    user, t = check_tournament_access(request, reg["tournament_id"])
    _delete_photo_from_s3(reg.get("photo_key", ""))
    delete_registration(registration_id)
    return HTMLResponse("", status_code=200, headers={"HX-Refresh": "true"})


# ── Registrations export (HTML + Excel) ──────────────────────────────────────
def _registrations_rows(tournament_id: str) -> List[Dict[str, Any]]:
    return list_registrations_by_tournament(tournament_id)


@app.get("/tournaments/{tournament_id}/registrations/export.html", response_class=HTMLResponse)
def registrations_export_html(request: Request, tournament_id: str):
    user, t = check_tournament_access(request, tournament_id)
    regs = _registrations_rows(tournament_id)
    return templates.TemplateResponse("registrations_export.html", {
        "request": request, "tournament": t, "registrations": regs,
    })


@app.get("/tournaments/{tournament_id}/registrations/export.xlsx")
def registrations_export_xlsx(request: Request, tournament_id: str):
    user, t = check_tournament_access(request, tournament_id)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on the server — run `pip install -r requirements.txt`")

    regs = _registrations_rows(tournament_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    headers = ["#", "Name", "Email", "Phone", "ITS", "Age", "Experience",
               "Format", "Team", "Partner", "Payment done", "Registered at"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, r in enumerate(regs, start=1):
        is_doubles = r.get("match_type") == "doubles"
        ws.append([
            idx,
            r.get("name", ""),
            r.get("email", ""),
            r.get("phone", ""),
            r.get("its", ""),
            int(r.get("age", 0) or 0),
            (r.get("experience", "") or "").title(),
            (r.get("match_type", "singles") or "singles").title(),
            r.get("team_name", "") if is_doubles else "",
            r.get("partner_name", "") if is_doubles else "",
            "Yes" if r.get("payment_done") else "No",
            r.get("created_at", ""),
        ])

    for col_idx, width in enumerate([5, 26, 30, 20, 16, 6, 14, 12, 24, 26, 14, 22], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in t.get("name", "tournament"))
    filename = f"{safe_name}-registrations.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/tournaments/{tournament_id}/participants", response_class=HTMLResponse)
def add_participant(request: Request, tournament_id: str,
                    name: str = Form(""), user_id: str = Form(""),
                    roster_player_id: str = Form("")):
    user, t = check_tournament_access(request, tournament_id)

    # If the user pasted comma/semicolon/newline-separated names into the single
    # "Add" input, route through the bulk logic instead so all get added.
    if not roster_player_id and not user_id and ("," in name or ";" in name or "\n" in name):
        return bulk_add_participants(request, tournament_id, names=name)

    display_name = name.strip()
    linked_user_id = user_id or ""
    if roster_player_id:
        rp = get_roster_player(roster_player_id)
        if rp:
            display_name = rp["name"]
            linked_user_id = rp.get("user_id", "") or linked_user_id
    if user_id and not display_name:
        u = get_user_by_id(user_id)
        if u:
            display_name = u["username"]
    if not display_name:
        raise HTTPException(400, "Name required")
    participants = t.get("participants", [])
    if any(p.get("name", "").lower() == display_name.lower() for p in participants):
        t2 = must_tournament(tournament_id)
        return templates.TemplateResponse("partials/tournament_body.html", {
            "request": request, "user": user, "tournament": t2,
            "roster": list_roster(),
            "flash": f"⚠️ {display_name} is already a participant",
        })
    pid = uuid.uuid4().hex[:8]
    participants.append({"id": pid, "name": display_name, "user_id": linked_user_id})
    update_tournament(tournament_id, "SET participants = :p", {":p": participants})
    t2 = must_tournament(tournament_id)
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": f"✅ Added {display_name}",
    })


@app.post("/tournaments/{tournament_id}/participants/remove", response_class=HTMLResponse)
def remove_participant(request: Request, tournament_id: str, participant_id: str = Form(...)):
    user, t = check_tournament_access(request, tournament_id)
    participants = [p for p in t.get("participants", []) if p.get("id") != participant_id]
    update_tournament(tournament_id, "SET participants = :p", {":p": participants})
    t2 = must_tournament(tournament_id)
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": "Removed participant",
    })


@app.post("/tournaments/{tournament_id}/participants/bulk", response_class=HTMLResponse)
def bulk_add_participants(request: Request, tournament_id: str, names: str = Form(...)):
    """Add many participants at once. Accepts one name per line
    (commas also allowed). Duplicates (case-insensitive) are skipped."""
    user, t = check_tournament_access(request, tournament_id)
    # Split on newlines, commas, semicolons; strip empties.
    raw = names.replace(",", "\n").replace(";", "\n").splitlines()
    incoming = [n.strip() for n in raw if n.strip()]
    if not incoming:
        raise HTTPException(400, "Provide at least one name")

    participants = list(t.get("participants", []))
    existing_lower = {p.get("name", "").lower() for p in participants}
    added, skipped = [], []
    for nm in incoming:
        key = nm.lower()
        if key in existing_lower:
            skipped.append(nm)
            continue
        existing_lower.add(key)
        pid = uuid.uuid4().hex[:8]
        participants.append({"id": pid, "name": nm, "user_id": ""})
        added.append(nm)

    update_tournament(tournament_id, "SET participants = :p", {":p": participants})
    t2 = must_tournament(tournament_id)
    flash_bits = []
    if added:
        flash_bits.append(f"✅ Added {len(added)}: {', '.join(added[:5])}{'…' if len(added) > 5 else ''}")
    if skipped:
        flash_bits.append(f"⚠️ Skipped {len(skipped)} duplicate(s)")
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": " • ".join(flash_bits) or "No participants added",
    })


@app.post("/tournaments/{tournament_id}/rename", response_class=HTMLResponse)
def rename_tournament(request: Request, tournament_id: str, name: str = Form(...)):
    user, t = check_tournament_access(request, tournament_id)
    new_name = name.strip()
    if not new_name:
        raise HTTPException(400, "Tournament name cannot be empty")
    if len(new_name) > 120:
        raise HTTPException(400, "Tournament name too long (max 120 chars)")
    update_tournament(
        tournament_id,
        "SET #n = :n",
        {":n": new_name},
        expr_names={"#n": "name"},
    )
    # Return the refreshed header snippet so htmx can swap it in-place
    t2 = must_tournament(tournament_id)
    return templates.TemplateResponse("partials/tournament_header.html", {
        "request": request, "user": user, "tournament": t2,
    })


@app.post("/tournaments/{tournament_id}/registration-settings", response_class=HTMLResponse)
def update_registration_settings(request: Request, tournament_id: str,
                                 format: str = Form("doubles"),
                                 registration_start: str = Form(""),
                                 registration_end: str = Form("")):
    user, t = check_tournament_access(request, tournament_id)
    fmt = (format or "doubles").strip().lower()
    if fmt not in ("singles", "doubles"):
        raise HTTPException(400, "format must be 'singles' or 'doubles'")
    update_tournament(
        tournament_id,
        "SET #f = :f, registration_start = :rs, registration_end = :re",
        {":f": fmt, ":rs": registration_start.strip(), ":re": registration_end.strip()},
        expr_names={"#f": "format"},
    )
    return HTMLResponse("", status_code=200, headers={"HX-Refresh": "true"})


def _group_registrations_for_build(regs: List[Dict[str, Any]]):
    """Split a list of registrations into (pairs, lone_singles).
    - pairs: list of dicts {pair_id, team_name, primary, partner}
    - lone_singles: list of registrations with no partner link.
    """
    by_pair: Dict[str, List[Dict[str, Any]]] = {}
    lone: List[Dict[str, Any]] = []
    for r in regs:
        pid = r.get("pair_id") or ""
        if r.get("match_type") == "doubles" and pid:
            by_pair.setdefault(pid, []).append(r)
        else:
            lone.append(r)
    pairs: List[Dict[str, Any]] = []
    for pid, members in by_pair.items():
        if len(members) < 2:
            # Orphaned doubles row — treat the survivor as a lone single.
            lone.extend(members)
            continue
        # Deterministic primary/partner ordering by created_at.
        members.sort(key=lambda m: m.get("created_at", ""))
        primary, partner = members[0], members[1]
        pairs.append({
            "pair_id": pid,
            "team_name": (primary.get("team_name") or partner.get("team_name") or "").strip(),
            "primary": primary,
            "partner": partner,
        })
    return pairs, lone


@app.get("/tournaments/{tournament_id}/build-participants", response_class=HTMLResponse)
def build_participants_form(request: Request, tournament_id: str):
    """Admin builder page: pick format, pair up lone singles, then confirm."""
    user, t = check_tournament_access(request, tournament_id)
    if (t.get("status") or "registration") != "registration":
        # Tournament already built — send admin back to the tournament page.
        return RedirectResponse(f"/tournaments/{tournament_id}", status_code=303)
    all_regs = list_registrations_by_tournament(tournament_id)
    paid_regs = [r for r in all_regs if r.get("payment_done")]
    unpaid_count = len(all_regs) - len(paid_regs)
    paid_pairs, paid_lone_singles = _group_registrations_for_build(paid_regs)
    paid_lone_singles.sort(key=lambda r: r.get("created_at", ""))
    return templates.TemplateResponse("tournament_build.html", {
        "request": request, "user": user, "tournament": t,
        "paid_pairs": paid_pairs,
        "paid_lone_singles": paid_lone_singles,
        "unpaid_count": unpaid_count,
        "total_regs": len(all_regs),
    })


@app.post("/tournaments/{tournament_id}/build-participants", response_class=HTMLResponse)
async def build_participants_from_registrations(request: Request, tournament_id: str):
    """Turn registrations into tournament participants and flip status → active.
    Reads the format + optional lone-single pair choices from the builder form."""
    user, t = check_tournament_access(request, tournament_id)
    if (t.get("status") or "registration") != "registration":
        raise HTTPException(400, "This tournament has already been built")

    form = await request.form()
    fmt = (form.get("format") or "doubles").strip().lower()
    if fmt not in ("singles", "doubles"):
        raise HTTPException(400, "format must be 'singles' or 'doubles'")
    include_unpaid = (form.get("include_unpaid") or "") in ("on", "true", "1", "yes")

    all_regs = list_registrations_by_tournament(tournament_id)
    regs = all_regs if include_unpaid else [r for r in all_regs if r.get("payment_done")]
    if not regs:
        raise HTTPException(400, "No matching registrations to import.")

    pairs, lone_singles = _group_registrations_for_build(regs)
    reg_by_id = {r["registration_id"]: r for r in regs}

    participants = list(t.get("participants", []))
    existing_names = {p.get("name", "").strip().lower() for p in participants}

    def _add_singles_participant(reg: Dict[str, Any]) -> None:
        nm = (reg.get("name") or "").strip()
        if not nm or nm.lower() in existing_names:
            return
        participants.append({
            "id": uuid.uuid4().hex[:8],
            "name": nm,
            "user_id": "",
        })
        existing_names.add(nm.lower())

    def _add_doubles_participant(primary: Dict[str, Any], partner: Dict[str, Any],
                                 team_name: str, pair_id: str = "") -> None:
        """Add BOTH members of a doubles pair as individual participants, linked by pair_id.
        This keeps the participant pool = players (not teams), so the round form can
        pick 4 distinct people for a doubles match, while match_context() still
        resolves the team badge via the shared team_name on each registration."""
        pnm = (primary.get("name") or "").strip()
        qnm = (partner.get("name") or "").strip()
        if not pnm or not qnm:
            return
        shared_pair = pair_id or uuid.uuid4().hex
        team_label = (team_name or "").strip() or f"{pnm} & {qnm}"
        for me, other in ((primary, partner), (partner, primary)):
            nm = (me.get("name") or "").strip()
            if not nm or nm.lower() in existing_names:
                continue
            participants.append({
                "id": uuid.uuid4().hex[:8],
                "name": nm,
                "user_id": "",
                "pair_id": shared_pair,
                "partner_name": (other.get("name") or "").strip(),
                "team_name": team_label,
            })
            existing_names.add(nm.lower())

    if fmt == "singles":
        # Every registration → 1 participant. Pairs are split into individuals.
        for r in regs:
            _add_singles_participant(r)
    else:
        # DOUBLES: registered pairs auto-import, admin-selected lone-single pairs also import.
        for pair in pairs:
            _add_doubles_participant(
                pair["primary"], pair["partner"],
                team_name=pair.get("team_name", ""),
                pair_id=pair["pair_id"],
            )
        # Collect admin-formed pairs from the form (lone_pair_<reg_id> = <other_reg_id>).
        seen: set = set()
        for s in lone_singles:
            sid = s["registration_id"]
            if sid in seen:
                continue
            other_id = (form.get(f"lone_pair_{sid}") or "").strip()
            if not other_id or other_id == sid or other_id in seen:
                continue
            other = reg_by_id.get(other_id)
            if not other:
                continue
            seen.add(sid)
            seen.add(other_id)
            team_name = (form.get(f"lone_team_{sid}") or "").strip()
            _add_doubles_participant(s, other, team_name=team_name)
        # Any unpaired lone singles are dropped from a doubles tournament.

    if len(participants) == 0:
        raise HTTPException(400, "Nothing to import — everyone was skipped.")

    update_tournament(
        tournament_id,
        "SET participants = :p, #st = :s, #f = :f",
        {":p": participants, ":s": "active", ":f": fmt},
        expr_names={"#st": "status", "#f": "format"},
    )
    # Plain form submit → use a proper 303 redirect (also works for htmx, which
    # follows the redirect transparently).
    return RedirectResponse(f"/tournaments/{tournament_id}", status_code=303)


@app.post("/tournaments/{tournament_id}/rounds", response_class=HTMLResponse)
def add_round(request: Request, tournament_id: str, round_name: str = Form(...)):
    user, t = check_tournament_access(request, tournament_id)
    rounds = t.get("rounds", [])
    round_num = (max((int(r.get("round_num", 0)) for r in rounds), default=0)) + 1
    rounds.append({
        "round_num": round_num,
        "name": round_name.strip() or f"Round {round_num}",
        "matches": [],
    })
    update_tournament(tournament_id, "SET rounds = :r", {":r": rounds})
    t2 = must_tournament(tournament_id)
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": f"✅ Added {round_name}",
    })


@app.post("/tournaments/{tournament_id}/rounds/{round_num}/pairs", response_class=HTMLResponse)
def add_pair(request: Request, tournament_id: str, round_num: int,
             match_type: str = Form("singles"),
             participant_a: str = Form(...), participant_b: str = Form(...),
             participant_a2: str = Form(""), participant_b2: str = Form("")):
    user, t = check_tournament_access(request, tournament_id)
    participants = {p["id"]: p for p in t.get("participants", [])}
    if match_type not in ("singles", "doubles"):
        match_type = "singles"

    # Enforce advancement rule: from round 2 onwards, only winners of the
    # previous round may be picked (falls back to full roster if the previous
    # round has no decided winners yet, so byes/corrections still work).
    allowed_pool = _advancing_participants(t, int(round_num))
    allowed_ids = {p["id"] for p in allowed_pool}

    a = participants.get(participant_a)
    b = participants.get(participant_b)
    if not a or not b or a["id"] == b["id"]:
        raise HTTPException(400, "Select two different participants")
    if a["id"] not in allowed_ids or b["id"] not in allowed_ids:
        raise HTTPException(400, "Only winners of the previous round can be added to this round")

    a2 = b2 = None
    if match_type == "doubles":
        a2 = participants.get(participant_a2)
        b2 = participants.get(participant_b2)
        if not a2 or not b2:
            raise HTTPException(400, "Doubles requires 4 participants")
        if a2["id"] not in allowed_ids or b2["id"] not in allowed_ids:
            raise HTTPException(400, "Only winners of the previous round can be added to this round")
        chosen = {a["id"], a2["id"], b["id"], b2["id"]}
        if len(chosen) != 4:
            raise HTTPException(400, "All 4 doubles participants must be distinct")

    rounds = t.get("rounds", [])
    target_round = next((r for r in rounds if int(r.get("round_num", 0)) == int(round_num)), None)
    if not target_round:
        raise HTTPException(404, "Round not found")
    slot = (max((int(s.get("slot", 0)) for s in target_round.get("matches", [])), default=0)) + 1
    pair = {
        "slot": slot,
        "match_type": match_type,
        "a_name": a["name"], "a_id": a["id"],
        "b_name": b["name"], "b_id": b["id"],
        "match_id": "",
        "winner": "",
        "winner_name": "",
    }
    if match_type == "doubles":
        pair["a2_name"] = a2["name"]; pair["a2_id"] = a2["id"]
        pair["b2_name"] = b2["name"]; pair["b2_id"] = b2["id"]
    target_round.setdefault("matches", []).append(pair)
    update_tournament(tournament_id, "SET rounds = :r", {":r": rounds})
    t2 = must_tournament(tournament_id)
    label = (f"{a['name']}/{a2['name']} vs {b['name']}/{b2['name']}"
             if match_type == "doubles" else f"{a['name']} vs {b['name']}")
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": f"✅ Pairing added: {label}",
    })


def _find_participant_by_name(participants: List[Dict[str, Any]], name: str) -> Optional[Dict[str, Any]]:
    """Case-insensitive lookup among tournament participants."""
    key = name.strip().lower()
    for p in participants:
        if p.get("name", "").strip().lower() == key:
            return p
    return None


def _parse_pair_line(line: str) -> Optional[Dict[str, Any]]:
    """Parse a single bulk pair line. Returns dict with keys:
      match_type: "singles" | "doubles"
      a_names, b_names: list[str]  (len 1 for singles, len 2 for doubles)
    Accepted separators between the two sides: 'vs', 'v', '-', '—', '|'.
    Doubles partners on one side separated by '+', '&', or '/'.
    """
    s = line.strip()
    if not s or s.startswith("#"):
        return None
    # Split sides
    import re
    parts = re.split(r"\s+(?:vs\.?|v\.?|—|-|\|)\s+", s, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) != 2:
        return None
    left, right = parts[0].strip(), parts[1].strip()
    if not left or not right:
        return None

    def _side(txt: str) -> List[str]:
        return [n.strip() for n in re.split(r"\s*[+&/]\s*", txt) if n.strip()]

    a_names = _side(left)
    b_names = _side(right)
    if not a_names or not b_names:
        return None
    if len(a_names) not in (1, 2) or len(b_names) not in (1, 2):
        return None
    if len(a_names) != len(b_names):
        return None  # mixed singles/doubles on one line is not allowed
    return {
        "match_type": "doubles" if len(a_names) == 2 else "singles",
        "a_names": a_names,
        "b_names": b_names,
    }


@app.post("/tournaments/{tournament_id}/rounds/{round_num}/pairs/bulk", response_class=HTMLResponse)
def bulk_add_pairs(request: Request, tournament_id: str, round_num: int,
                   pairings: str = Form(...)):
    """Bulk-create pairings. One per line. Supports:
        Alice vs Bob                        (singles)
        Alice + Amy vs Bob + Ben           (doubles)
    Participants must already exist on the tournament (case-insensitive match).
    Unknown names and malformed lines are reported back and skipped.
    """
    user, t = check_tournament_access(request, tournament_id)
    rounds = t.get("rounds", [])
    target_round = next((r for r in rounds if int(r.get("round_num", 0)) == int(round_num)), None)
    if not target_round:
        raise HTTPException(404, "Round not found")

    participants = t.get("participants", [])
    if not participants:
        raise HTTPException(400, "Add participants before pairing")

    lines = [ln for ln in pairings.splitlines() if ln.strip()]
    if not lines:
        raise HTTPException(400, "Provide at least one pairing")

    matches_list = target_round.setdefault("matches", [])
    next_slot = (max((int(s.get("slot", 0)) for s in matches_list), default=0)) + 1
    added, skipped = [], []

    for raw in lines:
        parsed = _parse_pair_line(raw)
        if not parsed:
            skipped.append(f"⚠️ Bad line: {raw.strip()[:60]}")
            continue
        # Resolve names → participant records.
        try:
            a_side = [_find_participant_by_name(participants, n) for n in parsed["a_names"]]
            b_side = [_find_participant_by_name(participants, n) for n in parsed["b_names"]]
        except Exception:
            skipped.append(f"⚠️ Bad line: {raw.strip()[:60]}")
            continue
        if any(p is None for p in a_side + b_side):
            missing = [n for n, p in
                       zip(parsed["a_names"] + parsed["b_names"], a_side + b_side)
                       if p is None]
            skipped.append(f"❓ Unknown participant(s): {', '.join(missing)}")
            continue
        chosen_ids = {p["id"] for p in a_side + b_side}
        if len(chosen_ids) != len(a_side) + len(b_side):
            skipped.append(f"⚠️ Duplicate participant: {raw.strip()[:60]}")
            continue
        pair = {
            "slot": next_slot,
            "match_type": parsed["match_type"],
            "a_name": a_side[0]["name"], "a_id": a_side[0]["id"],
            "b_name": b_side[0]["name"], "b_id": b_side[0]["id"],
            "match_id": "",
            "winner": "",
            "winner_name": "",
        }
        if parsed["match_type"] == "doubles":
            pair["a2_name"] = a_side[1]["name"]; pair["a2_id"] = a_side[1]["id"]
            pair["b2_name"] = b_side[1]["name"]; pair["b2_id"] = b_side[1]["id"]
            added.append(f"{a_side[0]['name']}/{a_side[1]['name']} vs {b_side[0]['name']}/{b_side[1]['name']}")
        else:
            added.append(f"{a_side[0]['name']} vs {b_side[0]['name']}")
        matches_list.append(pair)
        next_slot += 1

    update_tournament(tournament_id, "SET rounds = :r", {":r": rounds})
    t2 = must_tournament(tournament_id)
    flash_parts = []
    if added:
        flash_parts.append(f"✅ Added {len(added)} pairing(s)")
    if skipped:
        flash_parts.append(" • ".join(skipped[:5]))
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": " • ".join(flash_parts) or "No pairings added",
    })


@app.post("/tournaments/{tournament_id}/rounds/{round_num}/pairs/{slot}/edit", response_class=HTMLResponse)
def edit_pair(request: Request, tournament_id: str, round_num: int, slot: int,
              match_type: str = Form("singles"),
              participant_a: str = Form(...), participant_b: str = Form(...),
              participant_a2: str = Form(""), participant_b2: str = Form("")):
    """Change who plays whom for an existing pair. Only allowed before a match
    has been started (once a match_id exists we don't rewrite history)."""
    user, t = check_tournament_access(request, tournament_id)
    rounds = t.get("rounds", [])
    target_round = next((r for r in rounds if int(r.get("round_num", 0)) == int(round_num)), None)
    if not target_round:
        raise HTTPException(404, "Round not found")
    pair = next((m for m in target_round.get("matches", []) if int(m.get("slot", 0)) == int(slot)), None)
    if not pair:
        raise HTTPException(404, "Pair not found")
    if pair.get("match_id"):
        raise HTTPException(400, "Match already started — delete it first to re-pair")

    if match_type not in ("singles", "doubles"):
        match_type = "singles"
    participants = {p["id"]: p for p in t.get("participants", [])}
    a = participants.get(participant_a); b = participants.get(participant_b)
    if not a or not b or a["id"] == b["id"]:
        raise HTTPException(400, "Pick two different participants")

    pair["match_type"] = match_type
    pair["a_name"] = a["name"]; pair["a_id"] = a["id"]
    pair["b_name"] = b["name"]; pair["b_id"] = b["id"]
    # Reset any prior doubles fields.
    for k in ("a2_name", "a2_id", "b2_name", "b2_id"):
        pair.pop(k, None)
    if match_type == "doubles":
        a2 = participants.get(participant_a2); b2 = participants.get(participant_b2)
        if not a2 or not b2:
            raise HTTPException(400, "Doubles requires 4 participants")
        if len({a["id"], a2["id"], b["id"], b2["id"]}) != 4:
            raise HTTPException(400, "All 4 doubles participants must be distinct")
        pair["a2_name"] = a2["name"]; pair["a2_id"] = a2["id"]
        pair["b2_name"] = b2["name"]; pair["b2_id"] = b2["id"]

    update_tournament(tournament_id, "SET rounds = :r", {":r": rounds})
    t2 = must_tournament(tournament_id)
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": "✅ Pairing updated",
    })


@app.post("/tournaments/{tournament_id}/rounds/{round_num}/pairs/{slot}/delete", response_class=HTMLResponse)
def delete_pair(request: Request, tournament_id: str, round_num: int, slot: int):
    """Remove a pair from the round. Refuses if the match has already been
    started — the scorer must first delete the match record."""
    user, t = check_tournament_access(request, tournament_id)
    rounds = t.get("rounds", [])
    target_round = next((r for r in rounds if int(r.get("round_num", 0)) == int(round_num)), None)
    if not target_round:
        raise HTTPException(404, "Round not found")
    pair = next((m for m in target_round.get("matches", []) if int(m.get("slot", 0)) == int(slot)), None)
    if not pair:
        raise HTTPException(404, "Pair not found")
    if pair.get("match_id"):
        raise HTTPException(400, "Match already started — delete the match first")
    target_round["matches"] = [
        m for m in target_round.get("matches", []) if int(m.get("slot", 0)) != int(slot)
    ]
    update_tournament(tournament_id, "SET rounds = :r", {":r": rounds})
    t2 = must_tournament(tournament_id)
    return templates.TemplateResponse("partials/tournament_body.html", {
        "request": request, "user": user, "tournament": t2,
        "roster": list_roster(),
        "flash": "🗑️ Pairing removed",
    })


@app.post("/tournaments/{tournament_id}/rounds/{round_num}/start/{slot}")
def start_pair_match(request: Request, tournament_id: str, round_num: int, slot: int):
    user, t = check_tournament_access(request, tournament_id)
    target_round = next((r for r in t.get("rounds", []) if int(r.get("round_num", 0)) == int(round_num)), None)
    if not target_round:
        raise HTTPException(404, "Round not found")
    pair = next((m for m in target_round.get("matches", []) if int(m.get("slot", 0)) == int(slot)), None)
    if not pair:
        raise HTTPException(404, "Pair not found")
    if pair.get("match_id"):
        return RedirectResponse(f"/matches/{pair['match_id']}", status_code=303)

    cfg = get_settings()
    match_id = uuid.uuid4().hex
    match_type = pair.get("match_type", "singles")
    if match_type == "doubles":
        name = (f"{t['name']} — {target_round['name']}: "
                f"{pair['a_name']}/{pair.get('a2_name','')} vs "
                f"{pair['b_name']}/{pair.get('b2_name','')}")
    else:
        name = f"{t['name']} — {target_round['name']}: {pair['a_name']} vs {pair['b_name']}"

    item = {
        "match_id": match_id,
        "name": name,
        "match_type": match_type,
        "player_a": pair["a_name"],
        "player_b": pair["b_name"],
        "best_of": int(t.get("best_of", cfg["default_best_of"])),
        "points_to_win": int(t.get("points_to_win", cfg["default_points_to_win"])),
        "service_interval": int(t.get("service_interval", cfg["service_interval"])),
        "deuce_interval": int(t.get("deuce_interval", cfg["deuce_interval"])),
        "deciding_side_change_at": int(cfg.get("deciding_side_change_at", 5)),
        # Snapshot hard-cap rule onto the tournament match too — mirrors
        # /matches so global settings changes don't retro-alter live matches.
        "hard_cap_enabled": bool(cfg.get("hard_cap_enabled", False)),
        "hard_cap_at": int(cfg.get("hard_cap_at", 15)),
        "first_server": "A",
        "created_at": now_ts(),
        "user_id": user["user_id"],
        "tournament_id": tournament_id,
        "round_num": int(round_num),
        "match_slot": int(slot),
        "stats_recorded": False,
    }
    if match_type == "doubles":
        item["player_a2"] = pair.get("a2_name", "")
        item["player_b2"] = pair.get("b2_name", "")
        item["first_server_side"] = 1
        item["first_receiver_side"] = 1
    put_match(item)
    pair["match_id"] = match_id
    update_tournament(tournament_id, "SET rounds = :r", {":r": t["rounds"]})
    return RedirectResponse(f"/matches/{match_id}", status_code=303)


# ── Public tournament dashboard ───────────────────────────────────────────────
def _build_rounds_view(t: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Compute the read-only rounds/matches view used by the public live boards.
    One row per bracket slot with names + live score/game state resolved."""
    rounds_view: List[Dict[str, Any]] = []
    for r in t.get("rounds", []) or []:
        matches_view: List[Dict[str, Any]] = []
        for pair in r.get("matches", []) or []:
            mid = pair.get("match_id") or ""
            row = {
                "slot": int(pair.get("slot", 0)),
                "match_type": pair.get("match_type", "singles"),
                "a_id": pair.get("a_id", ""),
                "b_id": pair.get("b_id", ""),
                "a2_id": pair.get("a2_id", ""),
                "b2_id": pair.get("b2_id", ""),
                "a_name": pair.get("a_name", ""),
                "b_name": pair.get("b_name", ""),
                "a2_name": pair.get("a2_name", ""),
                "b2_name": pair.get("b2_name", ""),
                "match_id": mid,
                "winner": pair.get("winner", ""),
                "winner_name": pair.get("winner_name", ""),
                "status": "pending",
                "a_score": 0, "b_score": 0,
                "a_games": 0, "b_games": 0,
                "current_game_num": 1,
                "is_deuce": False,
                "is_deciding_game": False,
            }
            if mid:
                m = get_match(mid)
                if m:
                    ev = list_events(mid)
                    st = compute_state(m, ev)
                    row["a_score"] = st.get("a_score", 0)
                    row["b_score"] = st.get("b_score", 0)
                    row["a_games"] = st.get("a_games", 0)
                    row["b_games"] = st.get("b_games", 0)
                    row["current_game_num"] = st.get("current_game_num", 1)
                    row["is_deuce"] = bool(st.get("is_deuce"))
                    row["is_deciding_game"] = bool(st.get("is_deciding_game"))
                    row["status"] = "finished" if st.get("match_winner") else "live"
            matches_view.append(row)
        rounds_view.append({
            "round_num": int(r.get("round_num", 0)),
            "name": r.get("name", ""),
            "matches": matches_view,
        })
    return rounds_view


@app.get("/live/tournaments/{tournament_id}", response_class=HTMLResponse)
def live_tournament(request: Request, tournament_id: str):
    """Public read-only board for a single tournament — every round + every
    match's live score inline. No auth required."""
    t = get_tournament(tournament_id)
    if not t:
        raise HTTPException(404, "Tournament not found")
    return templates.TemplateResponse("live_tournament.html", {
        "request": request,
        "tournament": t,
        "rounds_view": _build_rounds_view(t),
    })
